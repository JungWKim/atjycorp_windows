import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import time
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────
# 설정 값
# ───────────────────────────────────────────────────────
EMAIL       = "kakao_support"    # 로그인 이메일
PASSWORD    = "Support1111@"     # 로그인 비밀번호
GROUP_ID    = 300426718           # 대상 그룹 ID
RESPONSIBLE = "채진윤"            # RMA담당자 이름
OUTPUT_FILE = r"C:\Users\lionm\Downloads\RMA요청.xlsx"

# ───────────────────────────────────────────────────────
# 센터 매핑
# ───────────────────────────────────────────────────────
CENTER_MAP = {
    "AS1": "안산 IDC", "PG1": "판교 IDC",
    "AY1": "안양 IDC", "AY2": "안양 IDC",
    "HN1": "하남 IDC", "MD1": "목동 IDC",
    "GS1": "가산 IDC", "GS2": "가산 IDC"
}

# ───────────────────────────────────────────────────────
# 모델명 매핑
# ───────────────────────────────────────────────────────
MODEL_MAP = {
    "M120": "1288H V5", "M120H": "1288H V5",
    "M220": "2288H V5", "M220H": "2288H V5",
    "M520": "5288 V5"
}

def normalize_model(raw: str) -> str:
    key = re.sub(r"\s+", "", raw.upper())
    for k, v in MODEL_MAP.items():
        if key.startswith(k):
            return v
    return raw.strip()

# ───────────────────────────────────────────────────────
# 파트 카테고리 탐지 함수 (필요 시 파트 추가)
# ───────────────────────────────────────────────────────
KNOWN_PARTS = ["DISK", "MEMORY", "FAN", "CPU", "BOX", "NIC"]

def detect_part_category(raw: str) -> str:
    up = raw.upper()
    for p in KNOWN_PARTS:
        if p in up:
            return p
    return ""

# ───────────────────────────────────────────────────────
# 문자열 표시 너비 계산 (한글 2, 영문 등 1)
# ───────────────────────────────────────────────────────
def display_width(s: str) -> int:
    w = 0
    for ch in s:
        w += 2 if ord(ch) > 255 else 1
    return w

# ───────────────────────────────────────────────────────
# 로그인 및 세션 설정
# ───────────────────────────────────────────────────────
session = requests.Session()
login_url = "https://partners.agit.io/login"
headers = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
}
# 로그인 폼 가져오기
r = session.get(login_url, headers=headers)
r.raise_for_status()
soup = BeautifulSoup(r.text, "html.parser")
utf8_val = soup.find("input", {"name": "utf8"}).get("value", "")
auth_token = soup.find("input", {"name": "authenticity_token"}).get("value", "")
cont_val = soup.find("input", {"name": "continue"}).get("value", "")
if not auth_token:
    raise RuntimeError("❌ authenticity_token 파싱 실패")
# 로그인 요청
login_data = {"utf8": utf8_val, "authenticity_token": auth_token, "continue": cont_val,
              "email": EMAIL, "password": PASSWORD}
post = session.post(
    login_url, data=login_data,
    headers={"User-Agent": headers["User-Agent"],
             "Accept": headers["Accept"],
             "Content-Type": "application/x-www-form-urlencoded",
             "Referer": login_url},
    allow_redirects=False
)
if post.status_code != 302:
    raise RuntimeError(f"로그인 실패: HTTP {post.status_code}")

# 그룹 홈에서 uToken 추출
group_home = f"https://partners.agit.io/g/{GROUP_ID}/"
r2 = session.get(group_home, headers=headers)
r2.raise_for_status()
m = re.search(r'window\.initialState\s*=.*?uToken\s*:\s*"([^"]+)"', r2.text, re.DOTALL)
if not m:
    raise RuntimeError("uToken 추출 실패")
u_token = m.group(1)

# ───────────────────────────────────────────────────────
# RMA 게시글 수집 (status=0 요청, 1=진행, 2=완료 / 원하는 상태로 변경하여 사용)
# ───────────────────────────────────────────────────────
all_threads = []
url = f"https://partners.agit.io/api/groups/{GROUP_ID}/wall?status=0"
while True:
    resp = session.get(url, headers={
        "Accept": "application/json",
        "Authorization": f"Bearer {u_token}",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": group_home
    })
    resp.raise_for_status()
    data = resp.json()
    threads = data.get("threads", [])
    all_threads.extend(threads)
    nxt = data.get("next_url")
    if not nxt:
        break
    url = ("https://partners.agit.io" + nxt) if nxt.startswith("/") else nxt
    time.sleep(0.2)

# ───────────────────────────────────────────────────────
# 메시지 파싱
# ───────────────────────────────────────────────────────
def parse_message(raw: str) -> dict:
    parts = [p.strip() for p in raw.split("\n\n") if p.strip()]
    out = {}
    for part in parts:
        lines = part.splitlines()
        if len(lines) < 2:
            continue
        key = lines[0].split(". ", 1)[-1].strip()
        val = "\n".join(l.strip() for l in lines[1:] if l.strip())
        out[key] = val
    return out

# ───────────────────────────────────────────────────────
# 데이터 정리
# ───────────────────────────────────────────────────────
rows = []
for thr in all_threads:
    contents = thr.get("contents") or [{}]
    info = contents[0].get("content_info", {})
    if info.get("status") != 1:
        continue
    fld = parse_message(thr.get("message", ""))
    date = thr.get("created_at", "").split()[0]
    loc = fld.get("Location", "").split("-", 1)[0]
    raw_model = fld.get("Model", "")
    mapped_model = normalize_model(raw_model)
    part_raw = fld.get("장애품목 및 >파트넘버", "")

    rows.append({
        "센터 위치":       CENTER_MAP.get(loc, ""),
        "이슈 발생시간":    date,
        "HOSTNAME":       fld.get("Hostname", ""),
        "장애 서버모델":     mapped_model,
        "파트 구분":       detect_part_category(part_raw),
        "파트 모델명":      part_raw,
        "수량":           fld.get("수량", ""),
        "서버 시리얼 번호":  "",
        "파트 번호":        "",
        "처리상태":        "진행중",
        "회수여부":        "미회수",
        "RMA담당자":       RESPONSIBLE,
        "회수자":         "-"
    })

df = pd.DataFrame(rows)
df = df.sort_values("이슈 발생시간").reset_index(drop=True)

# ───────────────────────────────────────────────────────
# 엑셀 저장 및 포맷
# ───────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Sheet1")
    ws = writer.sheets["Sheet1"]

    # 열 너비 자동 맞춤
    for idx, col in enumerate(df.columns, start=1):
        max_width = display_width(col)
        for val in df[col].astype(str):
            w = display_width(val)
            if w > max_width:
                max_width = w
        ws.column_dimensions[get_column_letter(idx)].width = max_width + 2

    # 가운데 정렬, 래핑 해제
    for r in range(1, len(df) + 2):
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=False, vertical="center")

    # 수량 2 이상만 노란 강조
    qty_idx = df.columns.get_loc("수량") + 1
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for i, v in enumerate(df["수량"], start=2):
        try:
            if int(str(v).split()[0]) > 1:
                ws.cell(row=i, column=qty_idx).fill = yellow
        except:
            pass

print(f"✅ 엑셀 저장 완료 → {OUTPUT_FILE}")
